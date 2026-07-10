"""SQL compute tool over the agent's granted data files (#24).

Lets the agent run read-only SQL (DuckDB) over the CSV/TSV/Excel files in its granted folder, so
exact aggregations (sums, group-bys, joins) are computed by the engine — not by the LLM summing rows
by hand (slow, expensive, and error-prone). Each data file is exposed as a table named by its base
file name (each Excel sheet becomes its own table).

Security: only files under the granted folder (re-validated against the allow-list) are loaded, each
materialized into an in-memory table; then ``enable_external_access`` is disabled so the query can't
reach any other file or write to disk, and the SQL is additionally restricted to ``SELECT``/``WITH``.
Results are row- and cell-capped.
"""

import asyncio
import os
import re
from typing import Optional

import duckdb
from langchain_core.tools import BaseTool, tool

from src.app.core.common.config import settings
from src.app.core.common.logging import logger
from src.app.core.sandbox.paths import is_within_allowed_roots

try:
    from openpyxl import load_workbook
except ImportError:  # openpyxl is optional; without it, xlsx files are simply not exposed to SQL
    load_workbook = None

_DATA_EXTS = {".csv", ".tsv"}  # delimited text, loaded via read_csv_auto
_XLSX_EXTS = {".xlsx"}  # Excel, loaded per-sheet via read_xlsx (native in DuckDB)
_QUERYABLE_EXTS = _DATA_EXTS | _XLSX_EXTS
_MAX_ROWS = 200
_MAX_CELL = 200
_IDENT_RE = re.compile(r"[^a-zA-Z0-9_]+")
# Read-only intent: the statement must START with SELECT or WITH. A keyword scan over the whole SQL
# is NOT used — it false-matches data (e.g. the month value 'set' = setembro matched the SET
# keyword). Security comes from: the query starts with SELECT/WITH, it is a single statement (no ';'
# stacking), and the connection has enable_external_access disabled over an ephemeral in-memory DB
# (any write would hit only throwaway memory; no file or external DB is reachable).
_READONLY_RE = re.compile(r"^\s*(with|select)\b", re.IGNORECASE)


def _table_name(path: str) -> str:
    """A safe SQL identifier from a file's base name (e.g. ``vendas_2025.csv`` → ``vendas_2025``)."""
    base = os.path.splitext(os.path.basename(path))[0]
    name = _IDENT_RE.sub("_", base).strip("_").lower()
    if not name or name[0].isdigit():
        name = f"t_{name}"
    return name


def _list_data_files(root_dir: str) -> list[str]:
    """Absolute, normalized paths of the queryable data files (CSV/TSV/XLSX) under ``root_dir``."""
    found: list[str] = []
    for r, _dirs, files in os.walk(root_dir):
        for name in files:
            if os.path.splitext(name)[1].lower() in _QUERYABLE_EXTS:
                found.append(os.path.normpath(os.path.join(r, name)))
    return sorted(found)


def _cell(value) -> str:
    """Render one result cell to a short string."""
    if value is None:
        return ""
    text = str(value)
    return text[:_MAX_CELL]


def _unique_name(base: str, taken: dict[str, str]) -> str:
    """A table name not already used (suffixing ``_2``, ``_3``… on collision)."""
    if base not in taken:
        return base
    i = 2
    while f"{base}_{i}" in taken:
        i += 1
    return f"{base}_{i}"


def _load_csv(con: duckdb.DuckDBPyConnection, path: str, mapping: dict[str, str]) -> None:
    """Materialize a CSV/TSV file as one table (types inferred by ``read_csv_auto``)."""
    tname = _unique_name(_table_name(path), mapping)
    try:
        con.execute(f'CREATE TABLE "{tname}" AS SELECT * FROM read_csv_auto(?, header=true)', [path])
        mapping[tname] = os.path.basename(path)
    except Exception:  # noqa: BLE001 - one bad file must not break the whole connection
        logger.exception("compute_load_failed", path=path)


def _load_xlsx(con: duckdb.DuckDBPyConnection, path: str, mapping: dict[str, str]) -> None:
    """Materialize each sheet of an Excel file as its own table (``read_xlsx`` per sheet).

    A single-sheet workbook keeps the file's base name; a multi-sheet workbook suffixes each table
    with the sheet name so every sheet is queryable. Skipped if openpyxl is unavailable.
    """
    if load_workbook is None:
        return
    fname = os.path.basename(path)
    try:
        workbook = load_workbook(path, read_only=True)
        sheets = workbook.sheetnames
        workbook.close()
    except Exception:  # noqa: BLE001 - a malformed workbook must not break the connection
        logger.exception("compute_xlsx_open_failed", path=path)
        return
    base = _table_name(path)
    for sheet in sheets:
        raw = base if len(sheets) == 1 else f"{base}_{_IDENT_RE.sub('_', sheet).strip('_').lower()}"
        tname = _unique_name(raw or base, mapping)
        try:
            con.execute(f'CREATE TABLE "{tname}" AS SELECT * FROM read_xlsx(?, sheet=?)', [path, sheet])
            mapping[tname] = f"{fname} (aba {sheet})" if len(sheets) > 1 else fname
        except Exception:  # noqa: BLE001 - an empty/odd sheet must not break the others
            logger.exception("compute_xlsx_sheet_failed", path=path, sheet=sheet)


def _build_connection(root_dir: str) -> tuple[duckdb.DuckDBPyConnection, dict[str, str]]:
    """In-memory DuckDB with each data file materialized as a table; external access then disabled."""
    con = duckdb.connect(":memory:")
    mapping: dict[str, str] = {}  # table name -> file (display) name
    for path in _list_data_files(root_dir):
        if settings.SANDBOX_ALLOWED_ROOTS and not is_within_allowed_roots(path, settings.SANDBOX_ALLOWED_ROOTS):
            continue
        if os.path.splitext(path)[1].lower() in _XLSX_EXTS:
            _load_xlsx(con, path, mapping)
        else:
            _load_csv(con, path, mapping)
    # Lock it down: no file reads/writes past this point. The user SQL only sees the tables above.
    con.execute("SET enable_external_access=false")
    return con, mapping


def _do_list(root_dir: str) -> str:
    """Describe the available data tables + columns (sync; run off the event loop)."""
    con, mapping = _build_connection(root_dir)
    try:
        if not mapping:
            return "Nenhum arquivo de dados (CSV/TSV/Excel) na pasta para consultar com SQL."
        lines = []
        for tname, fname in mapping.items():
            cols = con.execute(f'PRAGMA table_info("{tname}")').fetchall()
            colnames = ", ".join(c[1] for c in cols)
            count = con.execute(f'SELECT count(*) FROM "{tname}"').fetchone()[0]
            lines.append(f"- {tname}  (arquivo: {fname}, {count} linhas)  colunas: {colnames}")
        return "Tabelas de dados disponíveis (consulte com `consultar_dados`, SQL DuckDB):\n" + "\n".join(lines)
    finally:
        con.close()


def _do_query(root_dir: str, sql: str) -> str:
    """Run one read-only query and format the result (sync; run off the event loop)."""
    query = sql.strip().rstrip(";").strip()
    if ";" in query:
        return "Consulta rejeitada: envie apenas UMA instrução `SELECT`/`WITH` (sem ';')."
    if not _READONLY_RE.match(query):
        return "Consulta rejeitada: apenas `SELECT`/`WITH` (somente leitura) são permitidos."
    con, mapping = _build_connection(root_dir)
    try:
        try:
            cursor = con.execute(query)
        except Exception as e:
            available = ", ".join(mapping) or "(nenhuma)"
            return f"Erro na consulta: {str(e)[:300]}\nTabelas disponíveis: {available}. Veja `listar_dados`."
        columns = [d[0] for d in cursor.description]
        rows = cursor.fetchmany(_MAX_ROWS + 1)
        truncated = len(rows) > _MAX_ROWS
        rows = rows[:_MAX_ROWS]
        header = " | ".join(columns)
        body = "\n".join(" | ".join(_cell(v) for v in row) for row in rows)
        out = f"[proveniência: SQL sobre {', '.join(sorted(set(mapping.values()))) or '—'}]\n{header}\n{body}"
        if truncated:
            out += f"\n… (resultado truncado em {_MAX_ROWS} linhas)"
        return out
    finally:
        con.close()


def make_compute_tools(
    user_id: Optional[int],
    agent_id: Optional[int],
    root_dir: Optional[str],
    session_id: Optional[str] = None,
) -> list[BaseTool]:
    """Build the SQL compute tools bound to a session's granted folder. Empty without a folder."""
    if user_id is None or not root_dir:
        return []

    @tool
    async def listar_dados() -> str:
        """Lista os arquivos de dados (CSV/TSV/Excel) da pasta como tabelas SQL, com colunas e nº de linhas.

        Use ANTES de `consultar_dados` para saber quais tabelas e colunas existem (equivale ao
        `describe` para os arquivos da pasta). Cada planilha do Excel (`.xlsx`) vira uma tabela; se o
        arquivo tiver várias abas, cada aba é uma tabela (nome do arquivo + nome da aba).
        """
        return await asyncio.to_thread(_do_list, root_dir)

    @tool
    async def consultar_dados(sql: str) -> str:
        """Executa SQL de LEITURA (DuckDB) sobre os arquivos de dados da pasta e devolve o resultado EXATO.

        USE ISTO para QUALQUER cálculo sobre CSV/TSV/Excel — soma, contagem, média, ranking,
        agrupamento, cruzamento. **NUNCA some/agregue linhas na mão** (é lento, caro e dá erro): deixe
        o SQL calcular. As tabelas são os arquivos da pasta (veja `listar_dados`); só `SELECT`/`WITH`.
        Exemplo:
        `SELECT regiao, SUM(receita) AS receita, SUM(unidades) AS un FROM vendas_2025
         WHERE mes IN ('jul','ago','set') GROUP BY regiao ORDER BY receita DESC`
        """
        return await asyncio.to_thread(_do_query, root_dir, sql)

    return [listar_dados, consultar_dados]
