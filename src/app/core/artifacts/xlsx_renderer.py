"""Render a SpreadsheetSpec to a native Excel (.xlsx) workbook via openpyxl.

A ``.xlsx`` is an OOXML zip package, not text — writing it through ``write_file`` produces a file
Excel refuses to open. This renderer builds a real workbook: one worksheet per ``Sheet``, a styled
header row from the columns, then the data rows.
"""

import re
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.app.core.artifacts.spec import SpreadsheetSpec, Template

# Excel forbids these characters in a sheet name and caps the name at 31 chars.
_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_NAME = 31


def _safe_sheet_name(name: str, index: int, used: set[str]) -> str:
    """A valid, unique Excel sheet name derived from ``name`` (deduped case-insensitively)."""
    cleaned = _INVALID_SHEET.sub(" ", name or "").strip() or f"Planilha{index + 1}"
    candidate = cleaned[:_MAX_SHEET_NAME]
    n = 1
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = cleaned[: _MAX_SHEET_NAME - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _cell(value: Any) -> Any:
    """Coerce a cell to something openpyxl can write (primitives pass through, else stringify)."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def render_xlsx(spec: SpreadsheetSpec, path: str, template: Optional[Template] = None) -> str:
    """Write ``spec`` as a .xlsx workbook at ``path`` and return the path."""
    template = template or Template()
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default sheet; we add our own

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=template.primary_color)

    used: set[str] = set()
    for index, sheet in enumerate(spec.sheets or []):
        worksheet = workbook.create_sheet(title=_safe_sheet_name(sheet.name, index, used))
        if sheet.columns:
            worksheet.append([_cell(c) for c in sheet.columns])
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            worksheet.freeze_panes = "A2"  # keep the header visible while scrolling
        for row in sheet.rows:
            worksheet.append([_cell(c) for c in row])

    if not workbook.sheetnames:
        workbook.create_sheet(title="Planilha1")  # a workbook must have at least one sheet

    workbook.save(path)
    return path
