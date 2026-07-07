"""Integration tests for the Data Agent's artifact tool wiring (#18 → #20/#21).

The tool is the linchpin that activates the artifact → episodic-event → metrics/reflection chain.
These tests exercise it directly (no LLM): the tool is bound per session, it renders a real file,
and generating it records an ``artifact_generated`` event that reflection (#20) turns into a
learned preference — proving the pipeline is actually connected end to end.
"""

import os
import uuid
import zipfile

import pytest
from httpx import AsyncClient

from src.app.agents.data_agent.artifact_tools import _output_dir, make_artifact_tools

pytestmark = pytest.mark.asyncio


def _tool_named(session_id: str, name: str, *args, **kwargs):
    """Return the bound tool with ``name`` for a session (raises if absent)."""
    tools = make_artifact_tools(1, 7, session_id, *args, **kwargs)
    return next(t for t in tools if t.name == name)


class TestToolBinding:
    def test_absent_without_session_or_user(self):
        assert make_artifact_tools(user_id=1, agent_id=7, session_id=None) == []
        assert make_artifact_tools(user_id=None, agent_id=7, session_id="s") == []

    def test_present_with_context(self):
        tools = make_artifact_tools(user_id=1, agent_id=7, session_id="s")
        assert {t.name for t in tools} == {"gerar_artefato", "gerar_planilha"}

    def test_output_dir_prefers_writable_folder(self, tmp_path):
        # Writable folder → write there (user finds it); read-only or no folder → temp dir.
        assert _output_dir("s", str(tmp_path), True) == str(tmp_path)
        assert _output_dir("s", str(tmp_path), False) != str(tmp_path)
        assert _output_dir("s", None, True) != str(tmp_path)


class TestArtifactPipeline:
    def _tool(self, session_id: str):
        return make_artifact_tools(user_id=1, agent_id=7, session_id=session_id)[0]

    async def test_tool_parks_action_without_rendering(self, client: AsyncClient):
        from src.app.init import pending_action_repository

        tool = self._tool(f"sess-{uuid.uuid4()}")
        out = await tool.ainvoke(
            {
                "titulo": "Relatorio de Vendas",
                "formato": "docx",
                "secoes": [{"titulo": "Resumo", "itens": [{"texto": "Receita subiu 12%", "fonte": "tabela vendas"}]}],
            }
        )
        # Outward-facing: parked for confirmation, NOT rendered inline.
        assert "aguardando sua confirmação" in out
        pending = [a for a in await pending_action_repository.list_pending(1) if a.action_type == "export_artifact"]
        assert pending, "expected a parked export_artifact action"
        assert not os.path.isfile(pending[-1].payload["path"])  # nothing on disk yet

    async def test_confirm_renders_records_event_and_feeds_reflection(self, client: AsyncClient):
        from src.app.core.learning import run_reflection
        from src.app.core.session.event_repository import SessionEventRepository
        from src.app.init import hitl_service, pending_action_repository

        tool = self._tool(f"sess-{uuid.uuid4()}")
        await tool.ainvoke(
            {
                "titulo": "Fechamento",
                "formato": "docx",
                "secoes": [{"titulo": "KPIs", "itens": [{"texto": "Meta batida", "fonte": "tabela metas"}]}],
            }
        )
        action = [a for a in await pending_action_repository.list_pending(1) if a.action_type == "export_artifact"][-1]
        path = action.payload["path"]

        # Confirming runs the executor: the file is rendered and the event recorded.
        await hitl_service.confirm(action.id, 1)
        assert os.path.isfile(path)
        assert path.endswith(".docx")

        events = await SessionEventRepository().get_agent_events(1, 7)
        assert any(e.event_type == "artifact_generated" and e.payload.get("format") == "docx" for e in events)

        # Reflection (#20) now has data — the loop is connected end to end.
        profile = await run_reflection(1, 7)
        assert profile["preferred_output_format"] == "docx"

    async def test_confirmed_pptx_lands_in_writable_folder_and_is_valid(self, client: AsyncClient, tmp_path):
        # Regression: the deliverable must land in the granted (writable) folder as a REAL OOXML
        # file — not a text file the agent wrote elsewhere. PowerPoint rejects non-zip .pptx.
        from src.app.init import hitl_service, pending_action_repository

        tool = make_artifact_tools(1, 7, f"sess-{uuid.uuid4()}", str(tmp_path), writable_folder=True)[0]
        await tool.ainvoke(
            {"titulo": "Deck", "formato": "pptx", "secoes": [{"titulo": "S", "itens": [{"texto": "ok"}]}]}
        )
        action = [a for a in await pending_action_repository.list_pending(1) if a.action_type == "export_artifact"][-1]
        path = action.payload["path"]
        assert os.path.dirname(path) == str(tmp_path)  # inside the granted folder

        await hitl_service.confirm(action.id, 1)
        assert zipfile.is_zipfile(path)  # a genuine .pptx (zip/OOXML), openable by PowerPoint

    async def test_invalid_format_is_rejected_without_writing(self, client: AsyncClient):
        tool = self._tool("sess-artifact-2")
        out = await tool.ainvoke(
            {"titulo": "X", "formato": "pdf", "secoes": [{"titulo": "S", "itens": [{"texto": "a"}]}]}
        )
        assert "Formato inválido" in out

    async def test_empty_sections_rejected(self, client: AsyncClient):
        tool = self._tool("sess-artifact-3")
        out = await tool.ainvoke({"titulo": "X", "formato": "docx", "secoes": []})
        assert "Nada a gerar" in out


class TestSpreadsheetPipeline:
    def _tool(self, session_id: str, *args, **kwargs):
        return _tool_named(session_id, "gerar_planilha", *args, **kwargs)

    async def test_tool_parks_spreadsheet_without_rendering(self, client: AsyncClient):
        from src.app.init import pending_action_repository

        tool = self._tool(f"sess-{uuid.uuid4()}")
        out = await tool.ainvoke(
            {
                "titulo": "Vendas 2024",
                "planilhas": [
                    {"nome": "Resumo", "colunas": ["Mês", "Total"], "linhas": [["Jan", 1000], ["Fev", 1500]]}
                ],
            }
        )
        assert "aguardando sua confirmação" in out
        action = [a for a in await pending_action_repository.list_pending(1) if a.action_type == "export_artifact"][-1]
        assert action.payload["kind"] == "spreadsheet"
        assert action.payload["fmt"] == "xlsx"
        assert action.payload["path"].endswith(".xlsx")
        assert not os.path.isfile(action.payload["path"])  # nothing on disk yet

    async def test_confirm_renders_valid_xlsx_with_data(self, client: AsyncClient, tmp_path):
        from openpyxl import load_workbook

        from src.app.init import hitl_service, pending_action_repository

        tool = self._tool(f"sess-{uuid.uuid4()}", str(tmp_path), writable_folder=True)
        await tool.ainvoke(
            {
                "titulo": "Relatorio",
                "planilhas": [{"nome": "Dados", "colunas": ["Produto", "Qtd"], "linhas": [["A", 3], ["B", 7]]}],
            }
        )
        action = [a for a in await pending_action_repository.list_pending(1) if a.action_type == "export_artifact"][-1]
        path = action.payload["path"]
        assert os.path.dirname(path) == str(tmp_path)

        await hitl_service.confirm(action.id, 1)
        # A genuine .xlsx (zip/OOXML) that openpyxl — and Excel — can open.
        assert zipfile.is_zipfile(path)
        wb = load_workbook(path)
        ws = wb["Dados"]
        assert [c.value for c in ws[1]] == ["Produto", "Qtd"]  # header row
        assert [c.value for c in ws[2]] == ["A", 3]
        assert [c.value for c in ws[3]] == ["B", 7]

    async def test_empty_planilhas_rejected(self, client: AsyncClient):
        tool = self._tool("sess-sheet-empty")
        out = await tool.ainvoke({"titulo": "X", "planilhas": []})
        assert "Nada a gerar" in out
